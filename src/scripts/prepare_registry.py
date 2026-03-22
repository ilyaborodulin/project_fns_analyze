from openpyxl import load_workbook
import pandas as pd

wb = load_workbook("data/Реестр.xlsx", read_only=True)
ws = wb["Лист1"]
ws.reset_dimensions()

row_iter = ws.iter_rows(min_row=3, min_col=1, max_col=23, values_only=True)

columns = next(row_iter)
filtered_rows = []

for row in row_iter:
    if row[2] == "Юридическое лицо":
        category = row[3]
        activity = str(row[6]) if row[6] is not None else ""

        if category in ["Микропредприятие", "Малое предприятие", "Среднее предприятие"] and activity.startswith("41.20"):
            filtered_rows.append(row)

df_filtered = pd.DataFrame(filtered_rows, columns=columns)

print("После фильтрации:", df_filtered.shape)
print("Пропуски в ИНН:", df_filtered["ИНН"].isna().sum())
print("Пропуски в ОГРН:", df_filtered["ОГРН"].isna().sum())
print("Пропуски в Регион:", df_filtered["Регион"].isna().sum())

df_filtered["ИНН"] = df_filtered["ИНН"].astype(str).str.strip()
df_filtered["ОГРН"] = df_filtered["ОГРН"].astype(str).str.strip()

print("Количество строк:", len(df_filtered))
print("Количество уникальных ИНН:", df_filtered["ИНН"].nunique())
print("Дубликаты по ИНН:", df_filtered.duplicated(subset=["ИНН"]).sum())

df_filtered.to_csv("data/registry_filtered.csv", index=False, encoding="utf-8-sig")
print("Файл сохранен: data/registry_filtered.csv")